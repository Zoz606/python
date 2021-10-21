import openpyxl

invFile = openpyxl.load_workbook("inventory.xlsx")
productList = invFile['Sheet1']

productsPerSupplier = {}
totalValuePerSupplier = {}
productsUnderTenInv = {}

for productRow in range(2, productList.max_row + 1):
    supplierName = productList.cell(productRow, 4).value
    inventory = productList.cell(productRow, 2).value
    price = productList.cell(productRow, 3).value
    producrNum = int(productList.cell(productRow, 1).value)
    inventoryPrice = productList.cell(productRow, 5)

    # Calculation number for products per supplier
    if supplierName in productsPerSupplier:
        currentNumProducts = productsPerSupplier.get(supplierName)
        productsPerSupplier[supplierName] = currentNumProducts + 1
    else:
        productsPerSupplier[supplierName] = 1

    # Calculation total value of inventory per supplier
    if supplierName in totalValuePerSupplier:
        currentTotalValue = totalValuePerSupplier.get(supplierName)
        totalValuePerSupplier[supplierName] = currentTotalValue + \
            (inventory * price)
    else:
        totalValuePerSupplier[supplierName] = inventory * price

    # Logic products with inventory less than 10
    if inventory < 10:
        productsUnderTenInv[producrNum] = int(inventory)

    # Add value for total inventory price
    inventoryPrice.value = inventory * price


print(productsPerSupplier)
print(totalValuePerSupplier)
print(productsUnderTenInv)


invFile.save("inventory with total value.xlsx")
